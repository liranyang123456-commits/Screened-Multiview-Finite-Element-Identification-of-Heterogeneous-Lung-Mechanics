"""Loader and integrity checks for the ÉTS controlled-indentation dataset.

Dataset:
    Dahmani, Petit, and Laporte, "Dataset of deformed ultrasound images and
    signals under controlled indentation", Borealis, V1.
    https://doi.org/10.5683/SP3/ASTGWY

The public workbook synchronizes every B-mode/RF frame with commanded probe
indentation and measured contact force. Its mechanical-characterization plots
report Young's moduli measured independently on ten cylindrical specimens with
a Bose ElectroForce 3200. Those plot annotations are represented explicitly
below because the workbook stores them as EMF images rather than cells.
"""
from __future__ import annotations

import hashlib
import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import numpy as np


DATASET_DOI = "https://doi.org/10.5683/SP3/ASTGWY"
LICENSE = "CC0-1.0"
ROOT = Path(__file__).resolve().parent
RAW_ROOT = ROOT / "raw"
DATA_ROOT = ROOT / "data"
WORKBOOK = RAW_ROOT / "Acquisition informations.xlsx"

EXPECTED_MD5 = {
    "Ultrasound Acquisitions on phantom.7z": "5934072a8915186dff77732de7516c84",
    "Acquisition informations.xlsx": "eea796a50619aaa569973960c6503257",
    "README.txt": "239306ffc84476f387840425651b9121",
}


@dataclass(frozen=True)
class MechanicalTruth:
    """Independent compression-test truth transcribed from workbook plots."""

    background_mean_kpa: float = 31.0
    background_sd_kpa: float = 1.87
    inclusion_mean_kpa: float = 8.71
    inclusion_sd_kpa: float = 3.8
    specimens_per_region: int = 10
    instrument: str = "Bose ElectroForce 3200"
    source: str = "Acquisition informations.xlsx, mechanical characterization"


@dataclass(frozen=True)
class SequenceRecord:
    acquisition: int
    frame: int
    indentation_mm: float
    voltage_v: float
    force_n: float
    image_path: str
    rf_path: str | None


def _md5(path: Path) -> str:
    digest = hashlib.md5()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _indexed_files(directory: Path, prefix: str, suffix: str) -> dict[int, Path]:
    files: dict[int, Path] = {}
    pattern = re.compile(rf"^{re.escape(prefix)}0*(\d+){re.escape(suffix)}$", re.I)
    for path in directory.glob(f"{prefix}*{suffix}"):
        match = pattern.match(path.name)
        if match:
            files[int(match.group(1))] = path
    return files


def load_protocol(
    workbook: Path = WORKBOOK,
    data_root: Path = DATA_ROOT,
) -> list[SequenceRecord]:
    """Parse all valid frame/indentation/force rows from the public workbook."""
    import openpyxl

    sheet = openpyxl.load_workbook(
        workbook, data_only=True, read_only=True
    )["Index, indentation, and force"]
    records: list[SequenceRecord] = []

    for acquisition in range(1, 7):
        voltage_col = 3 + 2 * (acquisition - 1)
        force_col = voltage_col + 1
        sequence_root = data_root / f"Acqui{acquisition}"
        images = _indexed_files(sequence_root / "US_Image", "frame", ".png")
        rf_files = _indexed_files(sequence_root / "RF", "RF", ".mat")

        for row in range(3, sheet.max_row + 1):
            frame = sheet.cell(row, 1).value
            indentation = sheet.cell(row, 2).value
            voltage = sheet.cell(row, voltage_col).value
            force = sheet.cell(row, force_col).value
            if None in (frame, indentation, voltage, force):
                continue
            index = int(frame)
            if index not in images:
                raise FileNotFoundError(
                    f"Acqui{acquisition} frame {index} is listed in the workbook "
                    "but its B-mode image is missing."
                )
            records.append(
                SequenceRecord(
                    acquisition=acquisition,
                    frame=index,
                    indentation_mm=float(indentation),
                    voltage_v=float(voltage),
                    force_n=float(force),
                    image_path=str(images[index].relative_to(ROOT)),
                    rf_path=(
                        str(rf_files[index].relative_to(ROOT))
                        if index in rf_files
                        else None
                    ),
                )
            )
    return records


def validate_dataset(
    raw_root: Path = RAW_ROOT,
    data_root: Path = DATA_ROOT,
) -> dict[str, Any]:
    """Validate checksums, sequence counts, synchronization, and units."""
    checksums: dict[str, str] = {}
    for filename, expected in EXPECTED_MD5.items():
        path = raw_root / filename
        if not path.exists():
            raise FileNotFoundError(path)
        actual = _md5(path)
        if actual != expected:
            raise ValueError(f"MD5 mismatch for {filename}: {actual} != {expected}")
        checksums[filename] = actual

    records = load_protocol(raw_root / WORKBOOK.name, data_root)
    counts = {
        acquisition: sum(r.acquisition == acquisition for r in records)
        for acquisition in range(1, 7)
    }
    expected_counts = {1: 36, 2: 35, 3: 36, 4: 36, 5: 36, 6: 33}
    if counts != expected_counts:
        raise ValueError(f"Unexpected sequence counts: {counts}")

    for acquisition in range(1, 7):
        subset = [r for r in records if r.acquisition == acquisition]
        indentation = np.asarray([r.indentation_mm for r in subset])
        if np.any(np.diff(indentation) <= 0):
            raise ValueError(f"Acqui{acquisition} indentation is not increasing")
        if not np.allclose(np.diff(indentation), 0.2, atol=1e-8):
            raise ValueError(f"Acqui{acquisition} indentation step is not 0.2 mm")

    return {
        "doi": DATASET_DOI,
        "license": LICENSE,
        "checksums": checksums,
        "records": len(records),
        "records_with_rf": sum(record.rf_path is not None for record in records),
        "frames_per_acquisition": counts,
        "indentation_unit": "mm",
        "force_unit": "N",
        "mechanical_truth": asdict(MechanicalTruth()),
    }


def write_manifest(path: Path = ROOT / "manifest.json") -> dict[str, Any]:
    manifest = validate_dataset()
    records = load_protocol()
    manifest["sequences"] = [
        {
            "acquisition": acquisition,
            "records": [asdict(r) for r in records if r.acquisition == acquisition],
        }
        for acquisition in range(1, 7)
    ]
    path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return manifest


class ETSIndentationDataset:
    """Frame-level dataset with synchronized B-mode, indentation, and force."""

    def __init__(self, acquisition: int | None = None, load_rf: bool = False):
        if acquisition is not None and acquisition not in range(1, 7):
            raise ValueError("acquisition must be in [1, 6]")
        records = load_protocol()
        self.records = [
            record
            for record in records
            if acquisition is None or record.acquisition == acquisition
        ]
        self.load_rf = load_rf
        self.truth = MechanicalTruth()

    def __len__(self) -> int:
        return len(self.records)

    def __getitem__(self, index: int) -> dict[str, Any]:
        from PIL import Image

        record = self.records[index]
        image = np.asarray(
            Image.open(ROOT / record.image_path).convert("L"), dtype=np.float32
        )
        sample: dict[str, Any] = {
            "image": image[None] / 255.0,
            "acquisition": record.acquisition,
            "frame": record.frame,
            "indentation_mm": record.indentation_mm,
            "force_n": record.force_n,
            "voltage_v": record.voltage_v,
            "truth": self.truth,
        }
        if self.load_rf:
            from scipy.io import loadmat

            if record.rf_path is None:
                raise FileNotFoundError(
                    f"No RF file is distributed for Acqui{record.acquisition} "
                    f"frame {record.frame}; use the B-mode image or another frame."
                )
            sample["rf"] = loadmat(ROOT / record.rf_path)
        return sample


if __name__ == "__main__":
    summary = write_manifest()
    print(json.dumps({key: value for key, value in summary.items() if key != "sequences"}, indent=2))
